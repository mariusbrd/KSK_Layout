"""Runtime input lineage for Excel exports.

The dashboard calculations operate on prepared DataFrames, but the user-facing
exports also need to document which raw Excel inputs fed that prepared state.
This module inspects only workbook metadata/header rows and is called lazily
while an Excel export is being built.
"""

from __future__ import annotations

import hashlib
import io
import os
import posixpath
import zipfile
from collections.abc import Mapping
from dataclasses import dataclass
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

import pandas as pd


INPUT_LINEAGE_SHEET_NAME = "Input_Lineage"


@dataclass(frozen=True)
class ExcelInputSpec:
    role: str
    default_filename: str
    loader_key: str
    upload_keys: tuple[str, ...]
    header: int = 0
    sheets: tuple[str, ...] | None = None


MAIN_INPUT_SPECS: tuple[ExcelInputSpec, ...] = (
    ExcelInputSpec("Mitarbeiter", "Mitarbeiter.xlsx", "mitarbeiter", ("Mitarbeiter",)),
    ExcelInputSpec("Planstellen", "Planstellen.XLSX", "planstellen", ("Planstellen",)),
    ExcelInputSpec("ATZ", "ATZ.xlsx", "atz", ("ATZ",)),
    ExcelInputSpec("Ausbildung", "Ausbildung.xlsx", "ausbildung", ("Ausbildung",)),
)

TVOED_SPEC = ExcelInputSpec(
    "TV\u00d6D",
    "TV\u00d6D.xlsx",
    "tvoed",
    ("TV\u00d6D", "TVOED", "TV\xc3\u2013D", "TVOD"),
    header=1,
)


def _state_get(session_state: Any, key: str, default: Any = None) -> Any:
    if session_state is None:
        return default
    if hasattr(session_state, "get"):
        try:
            return session_state.get(key, default)
        except Exception:
            return default
    try:
        return session_state[key]
    except Exception:
        return default


def _current_session_state() -> Any:
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx

        if get_script_run_ctx(suppress_warning=True) is None:
            return {}
    except Exception:
        pass

    try:
        import streamlit as st

        return st.session_state
    except Exception:
        return {}


def _normalize_lookup_key(value: str) -> str:
    text = str(value).upper()
    replacements = {
        "\u00c4": "AE",
        "\u00d6": "OE",
        "\u00dc": "UE",
        "\xc3\u2013": "OE",
        "\xc3\u0153": "UE",
        "\xc3\u201e": "AE",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return "".join(ch for ch in text if ch.isalnum())


def _lookup_mapping_value(mapping: Mapping[str, Any] | None, keys: tuple[str, ...]) -> tuple[str | None, Any]:
    if not mapping:
        return None, None
    normalized = {_normalize_lookup_key(key): key for key in mapping}
    for candidate in keys:
        actual_key = normalized.get(_normalize_lookup_key(candidate))
        if actual_key is not None:
            return actual_key, mapping[actual_key]
    return None, None


def _file_signature(path: str | os.PathLike[str]) -> str:
    stat = os.stat(path)
    modified = datetime.fromtimestamp(stat.st_mtime).replace(microsecond=0).isoformat()
    return f"mtime={modified}; size_bytes={stat.st_size}"


def _bytes_signature(data: bytes) -> str:
    return f"sha256={hashlib.sha256(data).hexdigest()}; size_bytes={len(data)}"


def _source_label(path_or_name: str | None, fallback: str) -> str:
    return str(path_or_name or fallback)


def _coerce_file_bytes(value: Any) -> bytes | None:
    if value is None:
        return None
    if isinstance(value, bytes):
        return value
    if isinstance(value, bytearray):
        return bytes(value)
    if hasattr(value, "getvalue"):
        return value.getvalue()
    if hasattr(value, "read"):
        pos = value.tell() if hasattr(value, "tell") else None
        data = value.read()
        if pos is not None and hasattr(value, "seek"):
            value.seek(pos)
        return data
    raise TypeError(f"Unsupported upload value type: {type(value)!r}")


def _xlsx_member_path(base_dir: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(base_dir, target))


def _xml_text(element: ElementTree.Element | None) -> str | None:
    if element is None:
        return None
    parts = [text for text in element.itertext() if text is not None]
    text = "".join(parts)
    return text if text != "" else None


def _read_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    try:
        root = ElementTree.fromstring(zf.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    return [_xml_text(item) or "" for item in root]


def _cell_value(cell: ElementTree.Element, shared_strings: list[str]) -> str | None:
    cell_type = cell.attrib.get("t", "")
    if cell_type == "inlineStr":
        return _xml_text(next((child for child in cell if child.tag.endswith("}is")), None))

    raw_value = _xml_text(next((child for child in cell if child.tag.endswith("}v")), None))
    if raw_value is None:
        return None
    if cell_type == "s":
        try:
            return shared_strings[int(raw_value)]
        except (IndexError, ValueError):
            return raw_value
    return raw_value


def _read_xlsx_columns_fast(
    source: str | bytes,
    *,
    header: int = 0,
    sheets: tuple[str, ...] | None = None,
) -> list[tuple[str, list[str], str]]:
    """Read only requested xlsx header rows via zip/xml.

    Returns tuples of (sheet_name, columns, error). Any parsing issue raises so
    callers can fall back to openpyxl/pandas without changing output semantics.
    """

    source_obj: str | io.BytesIO = io.BytesIO(source) if isinstance(source, bytes) else source
    with zipfile.ZipFile(source_obj) as zf:
        workbook_root = ElementTree.fromstring(zf.read("xl/workbook.xml"))
        rels_root = ElementTree.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_targets = {
            rel.attrib["Id"]: _xlsx_member_path("xl", rel.attrib["Target"])
            for rel in rels_root
            if "Id" in rel.attrib and "Target" in rel.attrib
        }
        shared_strings = _read_shared_strings(zf)

        sheet_by_name: dict[str, str] = {}
        for sheet in workbook_root.iter():
            if not sheet.tag.endswith("}sheet"):
                continue
            name = sheet.attrib.get("name", "")
            rel_id = (
                sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                or sheet.attrib.get("r:id")
            )
            if name and rel_id and rel_id in rel_targets:
                sheet_by_name[name] = rel_targets[rel_id]

        requested_sheets = list(sheets or sheet_by_name.keys())
        target_row = str(int(header) + 1)
        rows: list[tuple[str, list[str], str]] = []
        for sheet_name in requested_sheets:
            member_path = sheet_by_name.get(str(sheet_name))
            if not member_path:
                rows.append((str(sheet_name), [], "Worksheet not found"))
                continue

            worksheet_root = ElementTree.fromstring(zf.read(member_path))
            header_row = None
            for row in worksheet_root.iter():
                if row.tag.endswith("}row") and row.attrib.get("r") == target_row:
                    header_row = row
                    break
            columns = []
            if header_row is not None:
                for cell in header_row:
                    if not cell.tag.endswith("}c"):
                        continue
                    value = _cell_value(cell, shared_strings)
                    if value is not None:
                        columns.append(str(value))
            rows.append((str(sheet_name), columns, ""))
        return rows


@lru_cache(maxsize=64)
def _read_workbook_columns_cached(
    source: str | bytes,
    *,
    role: str,
    source_type: str,
    file_name: str,
    path: str = "",
    header: int = 0,
    sheets: tuple[str, ...] | None = None,
    signature: str = "",
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    excel_source: str | io.BytesIO
    excel_source = io.BytesIO(source) if isinstance(source, bytes) else source

    try:
        fast_rows = _read_xlsx_columns_fast(source, header=header, sheets=sheets)
        return [
            {
                "Input-Rolle": role,
                "Quelle-Typ": source_type,
                "Datei": file_name,
                "Pfad": path,
                "Sheet": sheet_name,
                "Header-Zeile": int(header) + 1,
                "Spaltenanzahl": len(columns),
                "Spalten": "; ".join(columns),
                "Dateisignatur": signature,
                "Ermittlungsstatus": "fehler" if error else "ok",
                "Hinweis": error,
            }
            for sheet_name, columns, error in fast_rows
        ]
    except Exception:
        pass

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(excel_source, read_only=True, data_only=True)
        try:
            sheet_names = list(sheets or workbook.sheetnames)
            for sheet_name in sheet_names:
                try:
                    worksheet = workbook[str(sheet_name)]
                    row_values = next(
                        worksheet.iter_rows(min_row=int(header) + 1, max_row=int(header) + 1, values_only=True),
                        (),
                    )
                    columns = [str(value) for value in row_values if value is not None]
                    status = "ok"
                    error = ""
                except Exception as exc:
                    columns = []
                    status = "fehler"
                    error = str(exc)

                rows.append(
                    {
                        "Input-Rolle": role,
                        "Quelle-Typ": source_type,
                        "Datei": file_name,
                        "Pfad": path,
                        "Sheet": str(sheet_name),
                        "Header-Zeile": int(header) + 1,
                        "Spaltenanzahl": len(columns),
                        "Spalten": "; ".join(columns),
                        "Dateisignatur": signature,
                        "Ermittlungsstatus": status,
                        "Hinweis": error,
                    }
                )
        finally:
            workbook.close()
        return rows
    except Exception:
        excel_source = io.BytesIO(source) if isinstance(source, bytes) else source

    try:
        xls = pd.ExcelFile(excel_source)
        sheet_names = list(sheets or xls.sheet_names)
        for sheet_name in sheet_names:
            try:
                header_df = pd.read_excel(xls, sheet_name=sheet_name, nrows=0, header=header)
                columns = [str(col) for col in header_df.columns]
                status = "ok"
                error = ""
            except Exception as exc:
                columns = []
                status = "fehler"
                error = str(exc)

            rows.append(
                {
                    "Input-Rolle": role,
                    "Quelle-Typ": source_type,
                    "Datei": file_name,
                    "Pfad": path,
                    "Sheet": str(sheet_name),
                    "Header-Zeile": int(header) + 1,
                    "Spaltenanzahl": len(columns),
                    "Spalten": "; ".join(columns),
                    "Dateisignatur": signature,
                    "Ermittlungsstatus": status,
                    "Hinweis": error,
                }
            )
    except Exception as exc:
        rows.append(
            {
                "Input-Rolle": role,
                "Quelle-Typ": source_type,
                "Datei": file_name,
                "Pfad": path,
                "Sheet": "",
                "Header-Zeile": int(header) + 1,
                "Spaltenanzahl": 0,
                "Spalten": "",
                "Dateisignatur": signature,
                "Ermittlungsstatus": "fehler",
                "Hinweis": str(exc),
            }
        )
    return rows


def _read_workbook_columns(
    source: str | bytes,
    *,
    role: str,
    source_type: str,
    file_name: str,
    path: str = "",
    header: int = 0,
    sheets: tuple[str, ...] | None = None,
    signature: str = "",
) -> list[dict[str, Any]]:
    return list(
        _read_workbook_columns_cached(
            source,
            role=role,
            source_type=source_type,
            file_name=file_name,
            path=path,
            header=header,
            sheets=sheets,
            signature=signature,
        )
    )


def _original_file_paths() -> dict[str, str]:
    from config.settings import BASE_DIR

    original_dir = Path(BASE_DIR).parent / "Original-Daten"
    return {
        "mitarbeiter": str(original_dir / "Mitarbeiter.xlsx"),
        "planstellen": str(original_dir / "Planstellen.XLSX"),
        "atz": str(original_dir / "ATZ.xlsx"),
        "ausbildung": str(original_dir / "Ausbildung.xlsx"),
    }


def _sample_data_path() -> str:
    from config.settings import BASE_DIR

    return str(Path(BASE_DIR) / "data" / "sample_data" / "hr_data.xlsx")


def _find_tvoed_path() -> str:
    from config.settings import BASE_DIR

    original_dir = Path(BASE_DIR).parent / "Original-Daten"
    possible_names = ("TV\u00d6D.xlsx", "TVOED.xlsx", "TV\u00d6D.XLSX", "TVOED.XLSX", "TVOED.XLS", "TVOE.xlsx")
    for name in possible_names:
        path = original_dir / name
        if path.exists():
            return str(path)
    if original_dir.exists():
        for path in original_dir.iterdir():
            normalized = _normalize_lookup_key(path.name)
            if normalized.startswith("TVO") and path.suffix.lower() in {".xlsx", ".xls"}:
                return str(path)
    return str(original_dir / "TV\u00d6D.xlsx")


def _uploaded_input_rows(
    spec: ExcelInputSpec,
    *,
    uploads: Mapping[str, Any],
    upload_metadata: Mapping[str, Any] | None,
) -> list[dict[str, Any]]:
    upload_key, uploaded = _lookup_mapping_value(uploads, spec.upload_keys)
    if upload_key is None:
        return [
            {
                "Input-Rolle": spec.role,
                "Quelle-Typ": "Upload fehlt",
                "Datei": spec.default_filename,
                "Pfad": "",
                "Sheet": "",
                "Header-Zeile": int(spec.header) + 1,
                "Spaltenanzahl": 0,
                "Spalten": "",
                "Dateisignatur": "",
                "Ermittlungsstatus": "nicht geladen",
                "Hinweis": "Mindestens ein anderer Upload ist aktiv; fuer diese Rolle wurde kein Upload gefunden.",
            }
        ]

    _, metadata = _lookup_mapping_value(upload_metadata, (upload_key, spec.role, *spec.upload_keys))
    metadata = metadata if isinstance(metadata, Mapping) else {}
    data = _coerce_file_bytes(uploaded) or b""
    file_name = _source_label(metadata.get("file_name") or getattr(uploaded, "name", None), spec.default_filename)
    return _read_workbook_columns(
        data,
        role=spec.role,
        source_type="Upload",
        file_name=file_name,
        header=spec.header,
        sheets=spec.sheets,
        signature=_bytes_signature(data),
    )


def _path_input_rows(spec: ExcelInputSpec, path: str | os.PathLike[str] | None, source_type: str) -> list[dict[str, Any]]:
    if not path or not os.path.exists(path):
        return [
            {
                "Input-Rolle": spec.role,
                "Quelle-Typ": source_type,
                "Datei": spec.default_filename,
                "Pfad": str(path or ""),
                "Sheet": "",
                "Header-Zeile": int(spec.header) + 1,
                "Spaltenanzahl": 0,
                "Spalten": "",
                "Dateisignatur": "",
                "Ermittlungsstatus": "nicht gefunden",
                "Hinweis": "Datei wurde nicht gefunden.",
            }
        ]

    absolute_path = os.path.abspath(path)
    return _read_workbook_columns(
        absolute_path,
        role=spec.role,
        source_type=source_type,
        file_name=os.path.basename(absolute_path),
        path=absolute_path,
        header=spec.header,
        sheets=spec.sheets,
        signature=_file_signature(absolute_path),
    )


def _main_workforce_input_rows(session_state: Any) -> list[dict[str, Any]]:
    uploads = _state_get(session_state, "global_uploads", {}) or {}
    upload_metadata = _state_get(session_state, "global_upload_metadata", {}) or {}

    rows: list[dict[str, Any]] = []
    if uploads:
        for spec in MAIN_INPUT_SPECS:
            rows.extend(_uploaded_input_rows(spec, uploads=uploads, upload_metadata=upload_metadata))
    else:
        ORIGINAL_FILES = _original_file_paths()
        missing_originals = [path for path in ORIGINAL_FILES.values() if not os.path.exists(path)]
        if not missing_originals:
            for spec in MAIN_INPUT_SPECS:
                rows.extend(_path_input_rows(spec, ORIGINAL_FILES.get(spec.loader_key), "Original-Daten"))
        else:
            sample_path = _sample_data_path()
            sample_spec = ExcelInputSpec(
                "Synthetischer Snapshot",
                os.path.basename(sample_path),
                "synthetic",
                tuple(),
                sheets=("snapshot_detail", "history_cube", "org_structure"),
            )
            rows.extend(_path_input_rows(sample_spec, sample_path, "Synthetische Testdaten"))

    return rows


def _tvoed_input_rows(session_state: Any) -> list[dict[str, Any]]:
    uploads = _state_get(session_state, "global_uploads", {}) or {}
    upload_metadata = _state_get(session_state, "global_upload_metadata", {}) or {}
    upload_key, _ = _lookup_mapping_value(uploads, TVOED_SPEC.upload_keys)
    if upload_key is not None:
        return _uploaded_input_rows(TVOED_SPEC, uploads=uploads, upload_metadata=upload_metadata)

    TVOED_FILE = _find_tvoed_path()
    if os.path.exists(TVOED_FILE):
        return _path_input_rows(TVOED_SPEC, TVOED_FILE, "Original-Daten")

    return [
        {
            "Input-Rolle": TVOED_SPEC.role,
            "Quelle-Typ": "Konfigurations-Fallback",
            "Datei": "",
            "Pfad": "",
            "Sheet": "",
            "Header-Zeile": 2,
            "Spaltenanzahl": 0,
            "Spalten": "",
            "Dateisignatur": "",
            "Ermittlungsstatus": "fallback",
            "Hinweis": "Keine TVOD-Exceldatei geladen; Verguetung nutzt BASE_SALARY/STEP_MULTIPLIER aus config.settings.",
        }
    ]


def _cluster_input_rows(session_state: Any) -> list[dict[str, Any]]:
    try:
        from dataloader.cluster_resolver import (
            MODE_SYNTHETIC,
            SUBTYPE_UI_UPLOAD_SESSION,
            get_active_cluster_source,
            get_active_cluster_source_from_session,
        )
    except Exception as exc:
        return [
            {
                "Input-Rolle": "Cluster",
                "Quelle-Typ": "Nicht ermittelbar",
                "Datei": "",
                "Pfad": "",
                "Sheet": "",
                "Header-Zeile": 1,
                "Spaltenanzahl": 0,
                "Spalten": "",
                "Dateisignatur": "",
                "Ermittlungsstatus": "fehler",
                "Hinweis": str(exc),
            }
        ]

    active_source = get_active_cluster_source_from_session(session_state)
    if active_source is None:
        active_source = get_active_cluster_source(session_state=session_state)
    if getattr(active_source, "mode", None) == MODE_SYNTHETIC:
        return [
            {
                "Input-Rolle": "Cluster",
                "Quelle-Typ": "Synthetischer/Fallback-Cluster",
                "Datei": getattr(active_source, "display_label", "") or "",
                "Pfad": "",
                "Sheet": "",
                "Header-Zeile": 1,
                "Spaltenanzahl": 0,
                "Spalten": "",
                "Dateisignatur": getattr(active_source, "source_signature", "") or "",
                "Ermittlungsstatus": "fallback",
                "Hinweis": "Keine aktive Cluster-Exceldatei; Standard-/Fallback-Zuordnung wird verwendet.",
            }
        ]

    if getattr(active_source, "subtype", None) == SUBTYPE_UI_UPLOAD_SESSION:
        debug_meta = getattr(active_source, "debug_meta", {}) or {}
        data = _coerce_file_bytes(debug_meta.get("source_bytes") or _state_get(session_state, "cluster_upload_active_bytes"))
        if data:
            return _read_workbook_columns(
                data,
                role="Cluster",
                source_type="Upload",
                file_name=getattr(active_source, "filename", None) or "Cluster-Upload.xlsx",
                sheets=("OrgUnits", "JobFamilies"),
                signature=_bytes_signature(data),
            )

    source_path = getattr(active_source, "source_path", None) or getattr(active_source, "persisted_local_path", None)
    return _path_input_rows(
        ExcelInputSpec("Cluster", getattr(active_source, "filename", None) or "OE_Cluster.xlsx", "cluster", tuple(), sheets=("OrgUnits", "JobFamilies")),
        source_path,
        "Clusterquelle",
    )


def build_input_lineage_dataframe(session_state: Any | None = None) -> pd.DataFrame:
    """Return raw Excel source files and exact workbook columns for the active dashboard data."""

    state = _current_session_state() if session_state is None else session_state
    rows: list[dict[str, Any]] = []
    rows.extend(_main_workforce_input_rows(state))
    rows.extend(_tvoed_input_rows(state))
    rows.extend(_cluster_input_rows(state))
    return pd.DataFrame(rows)


def summarize_input_lineage(input_df: pd.DataFrame) -> tuple[str, str]:
    """Return compact file and column summaries for the Lineage_Report sheet."""

    if input_df.empty:
        return "", ""

    file_parts = []
    column_parts = []
    for _, row in input_df.iterrows():
        role = str(row.get("Input-Rolle", "")).strip()
        file_name = str(row.get("Datei", "")).strip()
        source_type = str(row.get("Quelle-Typ", "")).strip()
        sheet = str(row.get("Sheet", "")).strip()
        columns = str(row.get("Spalten", "")).strip()

        source_label = file_name or source_type
        if role or source_label:
            file_parts.append(f"{role}: {source_label} [{source_type}]")
        if columns:
            sheet_label = f"{role}.{sheet}" if sheet else role
            column_parts.append(f"{sheet_label}: {columns}")

    return "; ".join(file_parts), " | ".join(column_parts)
