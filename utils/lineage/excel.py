"""Excel export helpers for dashboard lineage reports."""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from functools import lru_cache
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from utils.lineage.code_index import (
    DEFAULT_BASELINE_PATH,
    build_function_lineage_index_for_refs,
    compare_hash_baseline_for_refs,
    lineage_reference_signature_for_refs,
)
from utils.lineage.inputs import (
    INPUT_LINEAGE_SHEET_NAME,
    build_input_lineage_dataframe,
    summarize_input_lineage,
)
from utils.lineage.registry import LineageSpec, get_lineage_specs
from utils.lineage.transformations import (
    TRANSFORMATION_LINEAGE_SHEET_NAME,
    build_transformation_lineage_dataframe,
)


LINEAGE_SHEET_NAME = "Lineage_Report"
FORMULA_LEADING_CHARS = ("=", "+", "-", "@")


def _sanitize_excel_value(value: Any) -> Any:
    if isinstance(value, str) and value[:1] in FORMULA_LEADING_CHARS:
        return " " + value
    return value


def _join(values: Sequence[Any]) -> str:
    return "; ".join(str(value) for value in values if str(value).strip())


def _format_sources(spec: LineageSpec) -> str:
    parts = []
    for source in spec.sources:
        columns = f"({', '.join(source.columns)})" if source.columns else ""
        parts.append(f"{source.table}{columns}")
    return _join(parts)


def _format_parameters(spec: LineageSpec) -> str:
    parts = []
    for parameter in spec.parameters:
        required = "required" if parameter.required else "optional"
        parts.append(f"{parameter.name} [{parameter.source}, {required}]")
    return _join(parts)


def _lineage_code_summary(
    spec: LineageSpec,
    drift_by_key: Mapping[str, str],
    lineage_by_ref: Mapping[tuple[str, str], Any],
) -> tuple[str, str]:
    code_parts = []
    status_parts = []
    for ref in spec.calculations:
        lineage = lineage_by_ref.get((ref.file_glob, ref.function_name))
        if lineage is None:
            raise KeyError(f"Unresolved lineage code reference: {ref.file_glob}:{ref.function_name}")
        code_key = f"{lineage.file_path}:{lineage.function_name}"
        code_parts.append(f"{code_key}#{lineage.source_hash}")
        status_parts.append(f"{code_key}={drift_by_key.get(code_key, 'unbekannt')}")
    return _join(code_parts), _join(status_parts)


@lru_cache(maxsize=128)
def _static_lineage_rows(
    lineage_ids: tuple[str, ...],
    cache_signature: tuple[Any, ...],
) -> tuple[dict[str, Any], ...]:
    """Build context-independent lineage rows for selected dashboard elements."""

    _ = cache_signature
    specs = get_lineage_specs(lineage_ids)
    refs = tuple(ref for spec in specs for ref in spec.calculations)
    lineage_by_key = build_function_lineage_index_for_refs(refs)
    lineage_by_ref = {
        (lineage.file_glob, lineage.function_name): lineage
        for lineage in lineage_by_key.values()
    }
    drift_df = compare_hash_baseline_for_refs(refs)
    drift_by_key = dict(zip(drift_df["Code-Key"], drift_df["Status"], strict=False))
    rows: list[dict[str, Any]] = []
    for spec in specs:
        code_hashes, code_drift = _lineage_code_summary(spec, drift_by_key, lineage_by_ref)
        rows.append(
            {
                "Lineage-ID": spec.lineage_id,
                "Label": spec.label,
                "Seite": spec.page,
                "Bereich": spec.section,
                "Elementtyp": spec.display_type,
                "Einheit": spec.unit,
                "Datenbasis": spec.data_basis,
                "Quellen": _format_sources(spec),
                "Parameter": _format_parameters(spec),
                "Formel": spec.formula,
                "Filter": _join(spec.filters),
                "Datenfluss": spec.data_lineage,
                "Tests": _join(spec.tests),
                "Validierungsstatus": spec.validation_status,
                "Code-Hashes": code_hashes,
                "Code-Drift": code_drift,
                "Export-Kontext": "",
                "Hinweise": spec.notes,
            }
        )
    return tuple(rows)


def _baseline_file_signature() -> tuple[int, int]:
    stat = DEFAULT_BASELINE_PATH.stat()
    return stat.st_mtime_ns, stat.st_size


def _lineage_cache_signature(lineage_ids: Sequence[str]) -> tuple[Any, ...]:
    specs = get_lineage_specs(lineage_ids)
    refs = tuple(ref for spec in specs for ref in spec.calculations)
    return (lineage_reference_signature_for_refs(refs), _baseline_file_signature())


def build_lineage_export_dataframe(
    lineage_ids: Sequence[str],
    *,
    export_context: Mapping[str, Any] | None = None,
    input_lineage_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Build a workbook-ready lineage report for selected dashboard elements."""

    context_text = _join([f"{key}={value}" for key, value in (export_context or {}).items()])
    if input_lineage_df is None:
        input_lineage_df = build_input_lineage_dataframe()
    input_files, input_columns = summarize_input_lineage(input_lineage_df)
    rows = [
        dict(
            row,
            **{
                "Export-Kontext": context_text,
                "Excel-Input-Dateien": input_files,
                "Excel-Input-Spalten": input_columns,
            },
        )
        for row in _static_lineage_rows(tuple(lineage_ids), _lineage_cache_signature(lineage_ids))
    ]
    return pd.DataFrame(rows)


def _build_lineage_frames(
    lineage_ids: Sequence[str],
    *,
    export_context: Mapping[str, Any] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    input_lineage_df = build_input_lineage_dataframe().map(_sanitize_excel_value)
    transformation_lineage_df = build_transformation_lineage_dataframe(
        lineage_ids,
        export_context=dict(export_context or {}),
    ).map(_sanitize_excel_value)
    lineage_df = build_lineage_export_dataframe(
        lineage_ids,
        export_context=export_context,
        input_lineage_df=input_lineage_df,
    ).map(_sanitize_excel_value)
    return lineage_df, input_lineage_df, transformation_lineage_df


def _write_openpyxl_dataframe(
    workbook: Workbook,
    dataframe: pd.DataFrame,
    *,
    sheet_name: str,
    max_width: int = 80,
) -> None:
    safe_sheet_name = sheet_name[:31]
    if safe_sheet_name in workbook.sheetnames:
        del workbook[safe_sheet_name]
    worksheet = workbook.create_sheet(title=safe_sheet_name)
    worksheet.append(list(dataframe.columns))
    for _, row in dataframe.iterrows():
        worksheet.append(["" if pd.isna(value) else value for value in row.tolist()])

    for col_idx, col_cells in enumerate(worksheet.columns, start=1):
        max_len = max((len(str(cell.value or "")) for cell in col_cells), default=10)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, max_width)


def _write_xlsxwriter_dataframe(
    workbook: Any,
    dataframe: pd.DataFrame,
    *,
    sheet_name: str,
    header_format: Any,
    max_width: int = 80,
) -> None:
    worksheet = workbook.add_worksheet(sheet_name[:31])
    for col_idx, col_name in enumerate(dataframe.columns):
        worksheet.write(0, col_idx, col_name, header_format)
        max_len = len(str(col_name))
        for row_idx, value in enumerate(dataframe[col_name], start=1):
            display_value = "" if pd.isna(value) else value
            worksheet.write(row_idx, col_idx, display_value)
            max_len = max(max_len, len(str(display_value)))
        worksheet.set_column(col_idx, col_idx, min(max_len + 3, max_width))


def write_lineage_sheet(
    writer: pd.ExcelWriter,
    lineage_ids: Sequence[str] | None,
    *,
    export_context: Mapping[str, Any] | None = None,
    sheet_name: str = LINEAGE_SHEET_NAME,
) -> None:
    """Append a lineage report sheet to an open Excel writer."""

    if not lineage_ids:
        return

    lineage_df, input_lineage_df, transformation_lineage_df = _build_lineage_frames(lineage_ids, export_context=export_context)
    lineage_df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    if not input_lineage_df.empty:
        input_lineage_df.to_excel(writer, sheet_name=INPUT_LINEAGE_SHEET_NAME[:31], index=False)
    if not transformation_lineage_df.empty:
        transformation_lineage_df.to_excel(writer, sheet_name=TRANSFORMATION_LINEAGE_SHEET_NAME[:31], index=False)


def append_lineage_sheet_to_workbook(
    workbook: Workbook,
    lineage_ids: Sequence[str] | None,
    *,
    export_context: Mapping[str, Any] | None = None,
    sheet_name: str = LINEAGE_SHEET_NAME,
) -> None:
    """Append a lineage report sheet to an existing openpyxl workbook."""

    if not lineage_ids:
        return

    lineage_df, input_lineage_df, transformation_lineage_df = _build_lineage_frames(lineage_ids, export_context=export_context)
    _write_openpyxl_dataframe(workbook, lineage_df, sheet_name=sheet_name, max_width=60)

    if not input_lineage_df.empty:
        _write_openpyxl_dataframe(workbook, input_lineage_df, sheet_name=INPUT_LINEAGE_SHEET_NAME, max_width=80)
    if not transformation_lineage_df.empty:
        _write_openpyxl_dataframe(
            workbook,
            transformation_lineage_df,
            sheet_name=TRANSFORMATION_LINEAGE_SHEET_NAME,
            max_width=100,
        )


def add_lineage_worksheet(
    workbook: Any,
    lineage_ids: Sequence[str] | None,
    *,
    export_context: Mapping[str, Any] | None = None,
    sheet_name: str = LINEAGE_SHEET_NAME,
) -> None:
    """Append a lineage report sheet to a direct xlsxwriter workbook."""

    if not lineage_ids:
        return

    lineage_df, input_lineage_df, transformation_lineage_df = _build_lineage_frames(lineage_ids, export_context=export_context)
    header_format = workbook.add_format({
        "bold": True,
        "bg_color": "#0088DE",
        "font_color": "#FFFFFF",
        "border": 1,
    })
    _write_xlsxwriter_dataframe(workbook, lineage_df, sheet_name=sheet_name, header_format=header_format, max_width=60)

    if not input_lineage_df.empty:
        _write_xlsxwriter_dataframe(
            workbook,
            input_lineage_df,
            sheet_name=INPUT_LINEAGE_SHEET_NAME,
            header_format=header_format,
            max_width=80,
        )
    if not transformation_lineage_df.empty:
        _write_xlsxwriter_dataframe(
            workbook,
            transformation_lineage_df,
            sheet_name=TRANSFORMATION_LINEAGE_SHEET_NAME,
            header_format=header_format,
            max_width=100,
        )
