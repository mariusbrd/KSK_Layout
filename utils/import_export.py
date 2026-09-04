"""
Import/Export Modul fuer Job Family Definitionen.

Ermoeglicht den Import und Export von Job Family Definitionen
in verschiedenen Formaten (Excel, CSV, JSON).
"""

import os
import json
import pandas as pd
from typing import Dict, List, Optional, Tuple, BinaryIO, Union
from datetime import datetime
from io import BytesIO

from utils.lineage import append_lineage_sheet_to_workbook

# Optionaler Import fuer Excel-Funktionalitaet
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# =============================================================================
# KONSTANTEN
# =============================================================================

DEFAULT_COLORS = [
    "#0088DE", "#10b981", "#f59e0b", "#E94D3A", "#8b5cf6",
    "#ec4899", "#06b6d4", "#84cc16", "#eab308", "#a855f7"
]


# =============================================================================
# IMPORT FUNKTIONEN
# =============================================================================

def import_from_excel(
    file: Union[str, BinaryIO],
    family_column: str,
    pattern_column: Optional[str] = None,
    description_column: Optional[str] = None,
    qualification_column: Optional[str] = None,
    color_column: Optional[str] = None,
    sheet_name: Union[str, int] = 0
) -> Tuple[Dict, Dict]:
    """
    Importiert Job Family Definitionen aus einer Excel-Datei.

    Unterstuetzte Formate:
    1. Pattern-basiert:
       | Job Family | Pattern | Beschreibung | Min. Qualifikation |
       |------------|---------|--------------|-------------------|
       | IT         | *IT-*   | IT-Bereich   | Fachinformatiker  |
       | IT         | *Admin* | IT-Bereich   | Fachinformatiker  |

    2. Einfach (nur Family-Namen):
       | Job Family |
       |------------|
       | IT         |
       | Beratung   |

    Args:
        file: Dateipfad oder File-Objekt
        family_column: Name der Spalte mit Job Family Namen
        pattern_column: Name der Spalte mit Patterns (optional)
        description_column: Name der Spalte mit Beschreibungen (optional)
        qualification_column: Name der Spalte mit Qualifikationen (optional)
        color_column: Name der Spalte mit Farben (optional)
        sheet_name: Name oder Index des Sheets

    Returns:
        Tuple[definitions, import_report]

        definitions: v2-kompatible Job Family Definitionen
        import_report: {
            "total_rows": 50,
            "families_created": 10,
            "patterns_imported": 45,
            "warnings": ["..."]
        }
    """
    if not OPENPYXL_AVAILABLE:
        return {}, {"error": "openpyxl nicht installiert. Bitte 'pip install openpyxl' ausfuehren."}

    try:
        df = pd.read_excel(file, sheet_name=sheet_name)
    except Exception as e:
        return {}, {"error": f"Fehler beim Lesen der Excel-Datei: {e}"}

    return _process_import_dataframe(
        df, family_column, pattern_column,
        description_column, qualification_column, color_column
    )


def import_from_csv(
    file: Union[str, BinaryIO],
    family_column: str,
    pattern_column: Optional[str] = None,
    description_column: Optional[str] = None,
    qualification_column: Optional[str] = None,
    color_column: Optional[str] = None,
    delimiter: str = ";",
    encoding: str = "utf-8"
) -> Tuple[Dict, Dict]:
    """
    Importiert Job Family Definitionen aus einer CSV-Datei.

    Args:
        file: Dateipfad oder File-Objekt
        family_column: Name der Spalte mit Job Family Namen
        pattern_column: Name der Spalte mit Patterns (optional)
        description_column: Name der Spalte mit Beschreibungen (optional)
        qualification_column: Name der Spalte mit Qualifikationen (optional)
        color_column: Name der Spalte mit Farben (optional)
        delimiter: Trennzeichen (Standard: ;)
        encoding: Zeichenkodierung (Standard: utf-8)

    Returns:
        Tuple[definitions, import_report]
    """
    try:
        df = pd.read_csv(file, delimiter=delimiter, encoding=encoding)
    except UnicodeDecodeError:
        # Versuche mit anderer Kodierung
        try:
            df = pd.read_csv(file, delimiter=delimiter, encoding="latin-1")
        except Exception as e:
            return {}, {"error": f"Fehler beim Lesen der CSV-Datei: {e}"}
    except Exception as e:
        return {}, {"error": f"Fehler beim Lesen der CSV-Datei: {e}"}

    return _process_import_dataframe(
        df, family_column, pattern_column,
        description_column, qualification_column, color_column
    )


def import_from_json(file: Union[str, BinaryIO]) -> Tuple[Dict, Dict]:
    """
    Importiert Job Family Definitionen aus einer JSON-Datei.

    Unterstuetzte Formate:
    1. v2-Format (direkt kompatibel)
    2. v1-Format (wird konvertiert)
    3. Template-Format

    Args:
        file: Dateipfad oder File-Objekt

    Returns:
        Tuple[definitions, import_report]
    """
    try:
        if isinstance(file, str):
            with open(file, "r", encoding="utf-8") as f:
                data = json.load(f)
        else:
            content = file.read()
            if isinstance(content, bytes):
                content = content.decode("utf-8")
            data = json.loads(content)
    except json.JSONDecodeError as e:
        return {}, {"error": f"Ungueltige JSON-Datei: {e}"}
    except Exception as e:
        return {}, {"error": f"Fehler beim Lesen der JSON-Datei: {e}"}

    warnings = []
    families_count = 0
    patterns_count = 0

    # Pruefe Format
    if "jobfamilies" in data:
        # v2 oder Template-Format
        definitions = {
            "jobfamilies": {},
            "manual_overrides": data.get("manual_overrides", {}),
            "metadata": {
                "version": "2.0",
                "source": "json_import",
                "imported": datetime.now().isoformat(),
                "last_modified": datetime.now().isoformat()
            }
        }

        for family_name, family_def in data.get("jobfamilies", {}).items():
            definitions["jobfamilies"][family_name] = {
                "description": family_def.get("description", ""),
                "patterns": family_def.get("patterns", []),
                "min_qualification": family_def.get("min_qualification", ""),
                "color": family_def.get("color", DEFAULT_COLORS[families_count % len(DEFAULT_COLORS)]),
                "manual_assignments": family_def.get("manual_assignments", [])
            }
            families_count += 1
            patterns_count += len(family_def.get("patterns", []))

    elif isinstance(data, dict) and all(isinstance(v, list) for v in data.values()):
        # v1-Format: {"Family": ["pattern1", "pattern2"]}
        definitions = {
            "jobfamilies": {},
            "manual_overrides": {},
            "metadata": {
                "version": "2.0",
                "source": "json_import_v1_conversion",
                "imported": datetime.now().isoformat(),
                "last_modified": datetime.now().isoformat()
            }
        }

        warnings.append("v1-Format erkannt und konvertiert")

        for family_name, patterns in data.items():
            definitions["jobfamilies"][family_name] = {
                "description": "",
                "patterns": patterns if isinstance(patterns, list) else [],
                "min_qualification": "",
                "color": DEFAULT_COLORS[families_count % len(DEFAULT_COLORS)],
                "manual_assignments": []
            }
            families_count += 1
            patterns_count += len(patterns) if isinstance(patterns, list) else 0

    else:
        return {}, {"error": "Unbekanntes JSON-Format"}

    import_report = {
        "total_rows": families_count,
        "families_created": families_count,
        "patterns_imported": patterns_count,
        "warnings": warnings
    }

    return definitions, import_report


def _process_import_dataframe(
    df: pd.DataFrame,
    family_column: str,
    pattern_column: Optional[str],
    description_column: Optional[str],
    qualification_column: Optional[str],
    color_column: Optional[str]
) -> Tuple[Dict, Dict]:
    """
    Verarbeitet einen DataFrame und erstellt Job Family Definitionen.
    """
    warnings = []

    # Pruefe ob Family-Spalte existiert
    if family_column not in df.columns:
        return {}, {"error": f"Spalte '{family_column}' nicht gefunden"}

    definitions = {
        "jobfamilies": {},
        "manual_overrides": {},
        "metadata": {
            "version": "2.0",
            "source": "file_import",
            "imported": datetime.now().isoformat(),
            "last_modified": datetime.now().isoformat()
        }
    }

    families_created = 0
    patterns_imported = 0
    color_index = 0

    # Gruppiere nach Family
    for family_name in df[family_column].dropna().unique():
        family_name_str = str(family_name).strip()
        if not family_name_str:
            continue

        family_rows = df[df[family_column] == family_name]

        # Patterns sammeln
        patterns = []
        if pattern_column and pattern_column in df.columns:
            patterns = family_rows[pattern_column].dropna().unique().tolist()
            patterns = [str(p).strip() for p in patterns if str(p).strip()]

        # Beschreibung (erste nicht-leere)
        description = ""
        if description_column and description_column in df.columns:
            desc_values = family_rows[description_column].dropna()
            if len(desc_values) > 0:
                description = str(desc_values.iloc[0])

        # Qualifikation (erste nicht-leere)
        qualification = ""
        if qualification_column and qualification_column in df.columns:
            qual_values = family_rows[qualification_column].dropna()
            if len(qual_values) > 0:
                qualification = str(qual_values.iloc[0])

        # Farbe
        color = DEFAULT_COLORS[color_index % len(DEFAULT_COLORS)]
        if color_column and color_column in df.columns:
            color_values = family_rows[color_column].dropna()
            if len(color_values) > 0:
                color = str(color_values.iloc[0])
                if not color.startswith("#"):
                    color = DEFAULT_COLORS[color_index % len(DEFAULT_COLORS)]

        definitions["jobfamilies"][family_name_str] = {
            "description": description,
            "patterns": patterns,
            "min_qualification": qualification,
            "color": color,
            "manual_assignments": []
        }

        families_created += 1
        patterns_imported += len(patterns)
        color_index += 1

    # Warnungen
    if families_created == 0:
        warnings.append("Keine Job Families gefunden")

    if pattern_column and patterns_imported == 0:
        warnings.append("Keine Patterns gefunden - Sie koennen diese spaeter manuell hinzufuegen")

    import_report = {
        "total_rows": len(df),
        "families_created": families_created,
        "patterns_imported": patterns_imported,
        "warnings": warnings
    }

    return definitions, import_report


def import_mapping_list(
    file: Union[str, BinaryIO],
    job_column: str,
    family_column: str,
    file_type: str = "excel",
    delimiter: str = ";"
) -> Tuple[Dict, Dict]:
    """
    Importiert eine Liste von Job-zu-Family Zuordnungen.

    Format:
    | Planstelle / Job-Titel | Job Family |
    |------------------------|------------|
    | IT-Administrator       | IT         |
    | Kundenberater Privat   | Beratung   |

    Die Jobs werden als manual_assignments zur jeweiligen Family hinzugefuegt.

    Args:
        file: Dateipfad oder File-Objekt
        job_column: Name der Spalte mit Job-Titeln
        family_column: Name der Spalte mit Job Family Namen
        file_type: "excel" oder "csv"
        delimiter: Trennzeichen fuer CSV

    Returns:
        Tuple[definitions, import_report]
    """
    try:
        if file_type == "excel":
            df = pd.read_excel(file)
        else:
            df = pd.read_csv(file, delimiter=delimiter)
    except Exception as e:
        return {}, {"error": f"Fehler beim Lesen der Datei: {e}"}

    if job_column not in df.columns:
        return {}, {"error": f"Spalte '{job_column}' nicht gefunden"}
    if family_column not in df.columns:
        return {}, {"error": f"Spalte '{family_column}' nicht gefunden"}

    definitions = {
        "jobfamilies": {},
        "manual_overrides": {},
        "metadata": {
            "version": "2.0",
            "source": "mapping_list_import",
            "imported": datetime.now().isoformat(),
            "last_modified": datetime.now().isoformat()
        }
    }

    assignments_count = 0
    color_index = 0

    # Gruppiere nach Family
    for family_name in df[family_column].dropna().unique():
        family_name_str = str(family_name).strip()
        if not family_name_str:
            continue

        family_rows = df[df[family_column] == family_name]
        jobs = family_rows[job_column].dropna().unique().tolist()
        jobs = [str(j).strip() for j in jobs if str(j).strip()]

        if family_name_str not in definitions["jobfamilies"]:
            definitions["jobfamilies"][family_name_str] = {
                "description": "",
                "patterns": [],
                "min_qualification": "",
                "color": DEFAULT_COLORS[color_index % len(DEFAULT_COLORS)],
                "manual_assignments": jobs
            }
            color_index += 1
        else:
            definitions["jobfamilies"][family_name_str]["manual_assignments"].extend(jobs)

        assignments_count += len(jobs)

    import_report = {
        "total_rows": len(df),
        "families_created": len(definitions["jobfamilies"]),
        "assignments_imported": assignments_count,
        "warnings": []
    }

    return definitions, import_report


# =============================================================================
# EXPORT FUNKTIONEN
# =============================================================================

def export_to_excel(definitions: Dict, file_path: Optional[str] = None) -> Optional[BytesIO]:
    """
    Exportiert Job Family Definitionen als Excel-Datei.

    Sheets:
    1. "Uebersicht" - Zusammenfassung aller Families
    2. "Patterns" - Alle Patterns mit zugehoeriger Family
    3. "Manuelle Zuordnungen" - Liste aller manuellen Assignments
    4. "Metadata" - Version und Zeitstempel

    Args:
        definitions: Job Family Definitionen (v2-Format)
        file_path: Optionaler Speicherpfad (wenn None, wird BytesIO zurueckgegeben)

    Returns:
        BytesIO-Objekt wenn kein file_path angegeben, sonst None
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl nicht installiert")

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0088DE", end_color="0088DE", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # =========================================================================
    # Sheet 1: Uebersicht
    # =========================================================================
    ws_overview = wb.active
    ws_overview.title = "Uebersicht"

    headers = ["Job Family", "Beschreibung", "Patterns", "Manuelle Zuordnungen", "Min. Qualifikation", "Farbe"]
    for col, header in enumerate(headers, 1):
        cell = ws_overview.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    row = 2
    for family_name, family_def in definitions.get("jobfamilies", {}).items():
        ws_overview.cell(row=row, column=1, value=family_name)
        ws_overview.cell(row=row, column=2, value=family_def.get("description", ""))
        ws_overview.cell(row=row, column=3, value=len(family_def.get("patterns", [])))
        ws_overview.cell(row=row, column=4, value=len(family_def.get("manual_assignments", [])))
        ws_overview.cell(row=row, column=5, value=family_def.get("min_qualification", ""))
        ws_overview.cell(row=row, column=6, value=family_def.get("color", ""))
        row += 1

    # Spaltenbreiten anpassen
    ws_overview.column_dimensions["A"].width = 25
    ws_overview.column_dimensions["B"].width = 40
    ws_overview.column_dimensions["C"].width = 12
    ws_overview.column_dimensions["D"].width = 20
    ws_overview.column_dimensions["E"].width = 20
    ws_overview.column_dimensions["F"].width = 12

    # =========================================================================
    # Sheet 2: Patterns
    # =========================================================================
    ws_patterns = wb.create_sheet("Patterns")

    headers = ["Job Family", "Pattern"]
    for col, header in enumerate(headers, 1):
        cell = ws_patterns.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    row = 2
    for family_name, family_def in definitions.get("jobfamilies", {}).items():
        for pattern in family_def.get("patterns", []):
            ws_patterns.cell(row=row, column=1, value=family_name)
            ws_patterns.cell(row=row, column=2, value=pattern)
            row += 1

    ws_patterns.column_dimensions["A"].width = 25
    ws_patterns.column_dimensions["B"].width = 40

    # =========================================================================
    # Sheet 3: Manuelle Zuordnungen
    # =========================================================================
    ws_manual = wb.create_sheet("Manuelle Zuordnungen")

    headers = ["Job Family", "Job-Titel / Planstelle"]
    for col, header in enumerate(headers, 1):
        cell = ws_manual.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    row = 2
    for family_name, family_def in definitions.get("jobfamilies", {}).items():
        for assignment in family_def.get("manual_assignments", []):
            ws_manual.cell(row=row, column=1, value=family_name)
            ws_manual.cell(row=row, column=2, value=assignment)
            row += 1

    # Manual Overrides
    for job_title, family_name in definitions.get("manual_overrides", {}).items():
        ws_manual.cell(row=row, column=1, value=family_name)
        ws_manual.cell(row=row, column=2, value=f"{job_title} (Override)")
        row += 1

    ws_manual.column_dimensions["A"].width = 25
    ws_manual.column_dimensions["B"].width = 50

    # =========================================================================
    # Sheet 4: Metadata
    # =========================================================================
    ws_meta = wb.create_sheet("Metadata")

    metadata = definitions.get("metadata", {})
    meta_data = [
        ["Eigenschaft", "Wert"],
        ["Version", metadata.get("version", "2.0")],
        ["Erstellt", metadata.get("created", "")],
        ["Zuletzt geaendert", metadata.get("last_modified", "")],
        ["Quelle", metadata.get("source", "")],
        ["Export-Datum", datetime.now().isoformat()],
        ["Anzahl Families", len(definitions.get("jobfamilies", {}))],
        ["Anzahl Patterns", sum(len(f.get("patterns", [])) for f in definitions.get("jobfamilies", {}).values())],
    ]

    for row_idx, row_data in enumerate(meta_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_meta.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill

    ws_meta.column_dimensions["A"].width = 20
    ws_meta.column_dimensions["B"].width = 40

    append_lineage_sheet_to_workbook(
        wb,
        ["2-05"],
        export_context={
            "Exporttyp": "Jobfamily Definitionen",
            "Jobfamilies": len(definitions.get("jobfamilies", {})),
            "Patterns": sum(len(f.get("patterns", [])) for f in definitions.get("jobfamilies", {}).values()),
            "Manuelle Zuordnungen": sum(
                len(f.get("manual_assignments", [])) for f in definitions.get("jobfamilies", {}).values()
            ),
            "Overrides": len(definitions.get("manual_overrides", {})),
        },
    )

    # Speichern
    if file_path:
        wb.save(file_path)
        return None
    else:
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


def export_to_csv(
    definitions: Dict,
    file_path: Optional[str] = None,
    delimiter: str = ";"
) -> Optional[str]:
    """
    Exportiert Job Family Definitionen als CSV-Datei.

    Args:
        definitions: Job Family Definitionen (v2-Format)
        file_path: Optionaler Speicherpfad
        delimiter: Trennzeichen

    Returns:
        CSV-String wenn kein file_path angegeben, sonst None
    """
    rows = []

    for family_name, family_def in definitions.get("jobfamilies", {}).items():
        for pattern in family_def.get("patterns", []):
            rows.append({
                "Job Family": family_name,
                "Pattern": pattern,
                "Beschreibung": family_def.get("description", ""),
                "Min. Qualifikation": family_def.get("min_qualification", ""),
                "Farbe": family_def.get("color", "")
            })

        # Falls keine Patterns, trotzdem Family eintragen
        if not family_def.get("patterns"):
            rows.append({
                "Job Family": family_name,
                "Pattern": "",
                "Beschreibung": family_def.get("description", ""),
                "Min. Qualifikation": family_def.get("min_qualification", ""),
                "Farbe": family_def.get("color", "")
            })

    df = pd.DataFrame(rows)

    if file_path:
        df.to_csv(file_path, sep=delimiter, index=False, encoding="utf-8-sig")
        return None
    else:
        return df.to_csv(sep=delimiter, index=False, encoding="utf-8-sig")


def export_to_json(definitions: Dict, file_path: Optional[str] = None) -> Optional[str]:
    """
    Exportiert Job Family Definitionen als JSON-Datei.

    Args:
        definitions: Job Family Definitionen (v2-Format)
        file_path: Optionaler Speicherpfad

    Returns:
        JSON-String wenn kein file_path angegeben, sonst None
    """
    output_data = {
        "jobfamilies": definitions.get("jobfamilies", {}),
        "manual_overrides": definitions.get("manual_overrides", {}),
        "metadata": definitions.get("metadata", {
            "version": "2.0",
            "exported": datetime.now().isoformat()
        })
    }

    if file_path:
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(output_data, f, indent=2, ensure_ascii=False)
        return None
    else:
        return json.dumps(output_data, indent=2, ensure_ascii=False)


def export_mapping_report(
    df: pd.DataFrame,
    definitions: Dict,
    file_path: Optional[str] = None,
    planstelle_column: str = "Planstelle",
    jobfamily_column: str = "Jobfamily",
    match_type_column: str = "Jobfamily_match_type"
) -> Optional[BytesIO]:
    """
    Exportiert einen Mapping-Report als Excel-Datei.

    Der Report zeigt fuer jeden Job:
    - Planstelle / Job-Titel
    - Zugeordnete Job Family
    - Match-Typ (Pattern, Manual Assignment, Override, Unmapped)
    - Gematchtes Pattern (falls zutreffend)

    Args:
        df: DataFrame mit Job-Daten (muss Jobfamily-Spalte enthalten)
        definitions: Job Family Definitionen
        file_path: Optionaler Speicherpfad
        planstelle_column: Name der Planstellen-Spalte
        jobfamily_column: Name der Jobfamily-Spalte
        match_type_column: Name der Match-Type-Spalte

    Returns:
        BytesIO-Objekt wenn kein file_path angegeben, sonst None
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl nicht installiert")

    # Erstelle Report-DataFrame
    report_data = []

    for _, row in df.iterrows():
        planstelle = row.get(planstelle_column, "")
        jobfamily = row.get(jobfamily_column, "UNMAPPED")
        match_type = row.get(match_type_column, "unknown")

        # Finde gematchtes Pattern
        matched_pattern = ""
        if match_type == "pattern" and jobfamily in definitions.get("jobfamilies", {}):
            import fnmatch
            for pattern in definitions["jobfamilies"][jobfamily].get("patterns", []):
                if fnmatch.fnmatch(str(planstelle).lower(), pattern.lower()):
                    matched_pattern = pattern
                    break

        report_data.append({
            "Planstelle": planstelle,
            "Job Family": jobfamily,
            "Match-Typ": match_type,
            "Gematchtes Pattern": matched_pattern
        })

    report_df = pd.DataFrame(report_data)

    # Erstelle Excel-Workbook
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0088DE", end_color="0088DE", fill_type="solid")
    unmapped_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")

    # =========================================================================
    # Sheet 1: Mapping Details
    # =========================================================================
    ws_mapping = wb.active
    ws_mapping.title = "Mapping Details"

    # Header
    for col, header in enumerate(report_df.columns, 1):
        cell = ws_mapping.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Daten
    for row_idx, row_data in enumerate(report_df.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_mapping.cell(row=row_idx, column=col_idx, value=value)
            # Markiere UNMAPPED
            if col_idx == 2 and value == "UNMAPPED":
                for c in range(1, 5):
                    ws_mapping.cell(row=row_idx, column=c).fill = unmapped_fill

    # Spaltenbreiten
    ws_mapping.column_dimensions["A"].width = 40
    ws_mapping.column_dimensions["B"].width = 25
    ws_mapping.column_dimensions["C"].width = 20
    ws_mapping.column_dimensions["D"].width = 30

    # =========================================================================
    # Sheet 2: Statistiken
    # =========================================================================
    ws_stats = wb.create_sheet("Statistiken")

    # Match-Typ Statistiken
    match_type_stats = report_df["Match-Typ"].value_counts()
    total_jobs = len(report_df)
    mapped_jobs = len(report_df[report_df["Job Family"] != "UNMAPPED"])

    stats_data = [
        ["Statistik", "Wert"],
        ["Gesamtanzahl Jobs", total_jobs],
        ["Zugeordnete Jobs", mapped_jobs],
        ["Nicht zugeordnet", total_jobs - mapped_jobs],
        ["Mapping-Quote", f"{mapped_jobs / total_jobs * 100:.1f}%" if total_jobs > 0 else "0%"],
        ["", ""],
        ["Match-Typ", "Anzahl"],
    ]

    for match_type, count in match_type_stats.items():
        stats_data.append([match_type, count])

    stats_data.append(["", ""])
    stats_data.append(["Job Family", "Anzahl"])

    family_stats = report_df["Job Family"].value_counts()
    for family, count in family_stats.items():
        stats_data.append([family, count])

    for row_idx, row_data in enumerate(stats_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_stats.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1 or value == "Match-Typ" or value == "Job Family":
                cell.font = Font(bold=True)

    ws_stats.column_dimensions["A"].width = 25
    ws_stats.column_dimensions["B"].width = 15

    append_lineage_sheet_to_workbook(
        wb,
        ["2-06"],
        export_context={
            "Exporttyp": "Jobfamily Mapping Report",
            "Zeilen": total_jobs,
            "Zugeordnet": mapped_jobs,
            "Nicht zugeordnet": total_jobs - mapped_jobs,
            "Planstellen-Spalte": planstelle_column,
            "Jobfamily-Spalte": jobfamily_column,
            "Match-Typ-Spalte": match_type_column,
        },
    )

    # Speichern
    if file_path:
        wb.save(file_path)
        return None
    else:
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


# =============================================================================
# HILFSFUNKTIONEN
# =============================================================================

def detect_file_format(filename: str) -> str:
    """
    Erkennt das Dateiformat basierend auf der Endung.

    Args:
        filename: Dateiname

    Returns:
        "excel", "csv" oder "json"
    """
    filename_lower = filename.lower()

    if filename_lower.endswith((".xlsx", ".xls")):
        return "excel"
    elif filename_lower.endswith(".csv"):
        return "csv"
    elif filename_lower.endswith(".json"):
        return "json"
    else:
        return "unknown"


def get_column_suggestions(df: pd.DataFrame) -> Dict[str, List[str]]:
    """
    Schlaegt Spalten-Mappings basierend auf Spaltennamen vor.

    Args:
        df: DataFrame

    Returns:
        {
            "family": ["Job Family", "Familie", ...],
            "pattern": ["Pattern", "Muster", ...],
            "description": ["Beschreibung", ...],
            ...
        }
    """
    columns = df.columns.tolist()

    suggestions = {
        "family": [],
        "pattern": [],
        "description": [],
        "qualification": [],
        "color": [],
        "job": []
    }

    family_keywords = ["family", "familie", "job family", "jobfamily", "gruppe", "kategorie"]
    pattern_keywords = ["pattern", "muster", "regex", "wildcard"]
    description_keywords = ["beschreibung", "description", "desc", "kommentar"]
    qualification_keywords = ["qualifikation", "qualification", "ausbildung", "abschluss"]
    color_keywords = ["farbe", "color", "colour"]
    job_keywords = ["planstelle", "job", "titel", "position", "stelle"]

    for col in columns:
        col_lower = col.lower()

        for keyword in family_keywords:
            if keyword in col_lower:
                suggestions["family"].append(col)
                break

        for keyword in pattern_keywords:
            if keyword in col_lower:
                suggestions["pattern"].append(col)
                break

        for keyword in description_keywords:
            if keyword in col_lower:
                suggestions["description"].append(col)
                break

        for keyword in qualification_keywords:
            if keyword in col_lower:
                suggestions["qualification"].append(col)
                break

        for keyword in color_keywords:
            if keyword in col_lower:
                suggestions["color"].append(col)
                break

        for keyword in job_keywords:
            if keyword in col_lower:
                suggestions["job"].append(col)
                break

    return suggestions


def validate_import_file(
    file: Union[str, BinaryIO],
    file_type: str
) -> Tuple[bool, str, Optional[pd.DataFrame]]:
    """
    Validiert eine Import-Datei.

    Args:
        file: Dateipfad oder File-Objekt
        file_type: "excel", "csv" oder "json"

    Returns:
        Tuple[is_valid, message, preview_df]
    """
    try:
        if file_type == "excel":
            df = pd.read_excel(file)
            return True, f"{len(df)} Zeilen gefunden", df.head(10)

        elif file_type == "csv":
            # Versuche verschiedene Trennzeichen
            for delimiter in [";", ",", "\t"]:
                try:
                    if isinstance(file, str):
                        df = pd.read_csv(file, delimiter=delimiter)
                    else:
                        file.seek(0)
                        df = pd.read_csv(file, delimiter=delimiter)

                    if len(df.columns) > 1:
                        return True, f"{len(df)} Zeilen gefunden (Trennzeichen: '{delimiter}')", df.head(10)
                except Exception:
                    continue

            return False, "Konnte CSV nicht parsen", None

        elif file_type == "json":
            if isinstance(file, str):
                with open(file, "r", encoding="utf-8") as f:
                    data = json.load(f)
            else:
                content = file.read()
                if isinstance(content, bytes):
                    content = content.decode("utf-8")
                data = json.loads(content)

            if "jobfamilies" in data:
                families = len(data["jobfamilies"])
                return True, f"{families} Job Families gefunden", None
            elif isinstance(data, dict):
                return True, f"{len(data)} Eintraege gefunden", None
            else:
                return False, "Unbekanntes JSON-Format", None

        else:
            return False, f"Unbekanntes Dateiformat: {file_type}", None

    except Exception as e:
        return False, f"Fehler beim Lesen: {str(e)}", None
