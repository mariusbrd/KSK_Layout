"""
Tests für dataloader/data_integrity.py (Blocker B17 Folgearbeit).

Deckt ab: die Abweichungen, die zur B17-Diskrepanz (883 vs. 884 Köpfe) führten
(besetzte Planstelle ohne Mitarbeiter-Match), müssen beim Upload sichtbar
gemacht werden - inklusive sicherer Excel-Generierung (keine Formel-Injection
aus Upload-Daten).
"""

import zipfile
from io import BytesIO

import pandas as pd
import pytest

from dataloader.data_integrity import (
    build_integrity_report_excel,
    check_mitarbeiter_planstellen_integrity,
)


def _mitarbeiter_df(rows):
    return pd.DataFrame(rows)


def _planstellen_df(rows):
    return pd.DataFrame(rows)


def test_matching_files_report_as_clean():
    mitarbeiter = _mitarbeiter_df([
        {"PersNr": "000001", "Vorname": "Anna", "Nachname": "Muster"},
        {"PersNr": "000002", "Vorname": "Bernd", "Nachname": "Beispiel"},
    ])
    planstellen = _planstellen_df([
        {"Personalnummer": "000001", "Planstellennr": 1, "Kürzel OrgEinheit": "100"},
        {"Personalnummer": "000002", "Planstellennr": 2, "Kürzel OrgEinheit": "100"},
        {"Personalnummer": None, "Planstellennr": 3, "Kürzel OrgEinheit": "100"},  # vakant
    ])

    report = check_mitarbeiter_planstellen_integrity(mitarbeiter, planstellen)

    assert report.is_clean
    assert report.total_findings == 0
    assert report.mitarbeiter_row_count == 2
    assert report.planstellen_row_count == 3


def test_detects_besetzte_planstelle_ohne_mitarbeiter_match():
    """Das ist exakt die Ursache von Blocker B17: eine Personalnummer besetzt
    eine Planstelle, hat aber keinen Mitarbeiter.xlsx-Datensatz."""
    mitarbeiter = _mitarbeiter_df([
        {"PersNr": "000001", "Vorname": "Anna", "Nachname": "Muster"},
    ])
    planstellen = _planstellen_df([
        {"Personalnummer": "000001", "Planstellennr": 1, "Kürzel OrgEinheit": "100"},
        {"Personalnummer": "006098", "Planstellennr": 2, "Kürzel OrgEinheit": "900"},
    ])

    report = check_mitarbeiter_planstellen_integrity(mitarbeiter, planstellen)

    assert not report.is_clean
    orphan_check = next(c for c in report.checks if c.key == "planstelle_ohne_mitarbeiter")
    assert orphan_check.severity == "error"
    assert orphan_check.count == 1
    assert orphan_check.detail["Personalnummer"].tolist() == ["006098"]


def test_detects_mitarbeiter_ohne_planstelle():
    mitarbeiter = _mitarbeiter_df([
        {"PersNr": "000001", "Vorname": "Anna", "Nachname": "Muster"},
        {"PersNr": "000099", "Vorname": "Carla", "Nachname": "Ohnestelle"},
    ])
    planstellen = _planstellen_df([
        {"Personalnummer": "000001", "Planstellennr": 1, "Kürzel OrgEinheit": "100"},
    ])

    report = check_mitarbeiter_planstellen_integrity(mitarbeiter, planstellen)

    orphan_check = next(c for c in report.checks if c.key == "mitarbeiter_ohne_planstelle")
    assert orphan_check.severity == "warning"
    assert orphan_check.count == 1
    assert orphan_check.detail["PersNr"].tolist() == ["000099"]


def test_detects_duplicate_personalnummer_in_mitarbeiter():
    mitarbeiter = _mitarbeiter_df([
        {"PersNr": "000001", "Vorname": "Anna", "Nachname": "Muster"},
        {"PersNr": "000001", "Vorname": "Anna", "Nachname": "Muster-Dublette"},
    ])
    planstellen = _planstellen_df([
        {"Personalnummer": "000001", "Planstellennr": 1, "Kürzel OrgEinheit": "100"},
    ])

    report = check_mitarbeiter_planstellen_integrity(mitarbeiter, planstellen)

    dupe_check = next(c for c in report.checks if c.key == "doppelte_personalnummer_mitarbeiter")
    assert dupe_check.severity == "error"
    assert dupe_check.count == 2


def test_persnr_normalization_matches_float_and_padded_variants():
    """Planstellen liefert Personalnummer oft als float (6098.0), Mitarbeiter als
    gepaddeten String (006098) - beides muss als dieselbe Person erkannt werden."""
    mitarbeiter = _mitarbeiter_df([
        {"PersNr": "6098", "Vorname": "Dirk", "Nachname": "Beispiel"},
    ])
    planstellen = _planstellen_df([
        {"Personalnummer": 6098.0, "Planstellennr": 1, "Kürzel OrgEinheit": "100"},
    ])

    report = check_mitarbeiter_planstellen_integrity(mitarbeiter, planstellen)

    assert report.is_clean


def test_missing_expected_columns_reports_schema_error():
    mitarbeiter = pd.DataFrame({"NotPersNr": ["000001"]})
    planstellen = pd.DataFrame({"NotPersonalnummer": ["000001"]})

    report = check_mitarbeiter_planstellen_integrity(mitarbeiter, planstellen)

    schema_check = next(c for c in report.checks if c.key == "schema_unvollstaendig")
    assert schema_check.severity == "error"
    assert schema_check.count == 2


def test_excel_export_has_no_formula_injection_from_malicious_upload_data():
    """Nachname/Vorname stammen aus einer hochgeladenen Datei und könnten
    manipuliert sein (CSV/Excel-Formula-Injection, z. B. =HYPERLINK(...)).
    Die generierte Evaluations-Excel darf daraus keine echte Formel machen."""
    mitarbeiter = _mitarbeiter_df([
        {"PersNr": "000001", "Vorname": "=HYPERLINK(\"http://evil.example\")", "Nachname": "Muster"},
    ])
    planstellen = _planstellen_df([
        {"Personalnummer": "000099", "Planstellennr": 1, "Kürzel OrgEinheit": "100"},
    ])

    report = check_mitarbeiter_planstellen_integrity(mitarbeiter, planstellen)
    assert not report.is_clean

    xlsx_bytes = build_integrity_report_excel(report)

    with zipfile.ZipFile(BytesIO(xlsx_bytes)) as archive:
        sheet_names = [n for n in archive.namelist() if n.startswith("xl/worksheets/") and n.endswith(".xml")]
        assert sheet_names
        for name in sheet_names:
            xml = archive.read(name).decode("utf-8", errors="replace")
            assert "<f>" not in xml and "<f " not in xml

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(xlsx_bytes))
    orphan_sheet = next(
        wb[name] for name in wb.sheetnames
        if name.startswith("Mitarbeiter ohne zugeordnete")
    )
    header = [cell.value for cell in next(orphan_sheet.iter_rows(min_row=1, max_row=1))]
    vorname_col = header.index("Vorname") + 1
    data_row = [cell.value for cell in next(orphan_sheet.iter_rows(min_row=2, max_row=2))]
    assert data_row[vorname_col - 1] == ' =HYPERLINK("http://evil.example")'


def test_excel_export_is_clean_when_report_has_no_findings():
    mitarbeiter = _mitarbeiter_df([{"PersNr": "000001"}])
    planstellen = _planstellen_df([{"Personalnummer": "000001", "Planstellennr": 1}])

    report = check_mitarbeiter_planstellen_integrity(mitarbeiter, planstellen)
    assert report.is_clean

    xlsx_bytes = build_integrity_report_excel(report)
    assert len(xlsx_bytes) > 0
