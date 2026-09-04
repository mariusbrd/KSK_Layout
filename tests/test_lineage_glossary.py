from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from utils.lineage_glossary import (
    GLOSSARY_PARAMETER_SHEET_NAME,
    GLOSSARY_SHEET_NAME,
    GLOSSARY_SOURCE_SHEET_NAME,
    GLOSSARY_TRANSFORMATION_SHEET_NAME,
    build_complete_glossary_frames,
    build_complete_glossary_workbook_bytes,
    build_glossary_dataframe,
)
from utils.lineage_registry import iter_lineage_specs


def test_complete_glossary_contains_every_registered_lineage_id() -> None:
    expected_ids = {spec.lineage_id for spec in iter_lineage_specs()}

    glossary = build_glossary_dataframe()

    assert set(glossary["Metrik-ID"]) == expected_ids
    assert glossary["Metrik-ID"].is_unique
    assert not glossary["Transformationsschritte"].str.strip().eq("").any()


def test_complete_glossary_frames_have_expected_sheets_and_ids() -> None:
    expected_ids = {spec.lineage_id for spec in iter_lineage_specs()}

    frames = build_complete_glossary_frames()

    assert set(frames) == {
        GLOSSARY_SHEET_NAME,
        GLOSSARY_TRANSFORMATION_SHEET_NAME,
        GLOSSARY_SOURCE_SHEET_NAME,
        GLOSSARY_PARAMETER_SHEET_NAME,
    }
    for sheet_name, dataframe in frames.items():
        id_column = "Lineage-ID" if sheet_name == GLOSSARY_TRANSFORMATION_SHEET_NAME else "Metrik-ID"
        assert set(dataframe[id_column]) == expected_ids


def test_complete_glossary_workbook_is_excel_readable() -> None:
    payload = build_complete_glossary_workbook_bytes(export_context={"Exporttyp": "Test"})
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)

    assert GLOSSARY_SHEET_NAME in workbook.sheetnames
    assert GLOSSARY_TRANSFORMATION_SHEET_NAME in workbook.sheetnames
    assert GLOSSARY_SOURCE_SHEET_NAME in workbook.sheetnames
    assert GLOSSARY_PARAMETER_SHEET_NAME in workbook.sheetnames

    headers = [
        workbook[GLOSSARY_SHEET_NAME].cell(1, column_index).value
        for column_index in range(1, workbook[GLOSSARY_SHEET_NAME].max_column + 1)
    ]
    assert "Export-Kontext" in headers


def test_glossary_can_be_limited_to_selected_lineage_ids() -> None:
    glossary = build_glossary_dataframe(["1-02", "10-04"])

    assert glossary["Metrik-ID"].tolist() == ["1-02", "10-04"]
    assert "Verguetungs-Fit je Entgeltgruppen-Spanne berechnen" in glossary.loc[0, "Transformationsschritte"]
    assert "Abgaenge aus Simulation ableiten" in glossary.loc[1, "Transformationsschritte"]
