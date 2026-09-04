from __future__ import annotations

from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = ROOT.parent
GLOSSARY_DRAFT = PROJECT_ROOT / "docs" / "KPI_Glossar_Analyseflaechen_Entwurf.tsv"
WORKBOOK_PATH = PROJECT_ROOT / "docs" / "KPI_Glossar_und_Validierung.xlsx"
REVIEW_PATH = PROJECT_ROOT / "docs" / "KPI_Glossar_Fachreview_Analyseflaechen_2026-09-02.md"
TESTS_ROOT = ROOT / "tests"

REQUIRED_COLUMNS = [
    "Metrik-ID",
    "Seite / Bereich",
    "Darstellungstyp",
    "Kurzbeschreibung",
    "Einheit",
    "Datenbasis",
    "Code-Referenz",
    "Formel / Berechnungslogik",
    "Filterwirkung",
    "Data Lineage",
    "Testnachweis",
    "Validierungsstatus",
    "Offene Punkte",
]

EXPECTED_IDS = {
    "8-13",
    "8-14",
    "8-15",
    "8-16",
    "8-17",
    "9-13",
    "9-14",
    "9-15",
    "9-16",
    "9-17",
    "10-01",
    "10-02",
    "10-03",
    "10-04",
    "10-05",
    "10-06",
    "10-07",
    "11-01",
    "11-02",
    "11-03",
    "11-04",
    "11-05",
}

WORKBOOK_SHEETS_BY_PREFIX = {
    "8": "8_Jobgruppen_Analyse",
    "9": "9_Organisationseinheiten_Analys",
    "10": "10_OE_Analyse_Simulation",
    "11": "11_Jobgruppen_Analyse_Sim",
}


def _load_glossary_draft() -> pd.DataFrame:
    return pd.read_csv(GLOSSARY_DRAFT, sep="\t", dtype=str).fillna("")


def _available_test_functions() -> set[str]:
    functions: set[str] = set()
    for test_file in TESTS_ROOT.glob("test_*.py"):
        text = test_file.read_text(encoding="utf-8")
        functions.update(re.findall(r"^def (test_[A-Za-z0-9_]+)\(", text, flags=re.MULTILINE))
    return functions


def test_analysis_glossary_draft_has_required_schema_and_ids():
    df = _load_glossary_draft()

    assert df.columns.tolist() == REQUIRED_COLUMNS
    assert set(df["Metrik-ID"]) == EXPECTED_IDS
    assert df["Metrik-ID"].is_unique


def test_analysis_glossary_draft_required_fields_are_populated():
    df = _load_glossary_draft()
    required_content_columns = [
        "Seite / Bereich",
        "Darstellungstyp",
        "Kurzbeschreibung",
        "Einheit",
        "Datenbasis",
        "Code-Referenz",
        "Formel / Berechnungslogik",
        "Filterwirkung",
        "Data Lineage",
        "Testnachweis",
        "Validierungsstatus",
    ]

    for column in required_content_columns:
        empty_ids = df.loc[df[column].str.strip().eq(""), "Metrik-ID"].tolist()
        assert empty_ids == [], f"{column} is empty for {empty_ids}"


def test_analysis_glossary_draft_references_existing_tests():
    df = _load_glossary_draft()
    available_functions = _available_test_functions()
    referenced_functions: set[str] = set()

    for value in df["Testnachweis"]:
        referenced_functions.update(
            part.strip()
            for part in value.split(";")
            if part.strip().startswith("test_")
        )

    missing = sorted(referenced_functions - available_functions)
    assert missing == []


def test_analysis_glossary_workbook_contains_synced_rows():
    df = _load_glossary_draft()
    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)

    for prefix, sheet_name in WORKBOOK_SHEETS_BY_PREFIX.items():
        assert sheet_name in wb.sheetnames
        ws = wb[sheet_name]
        headers = [ws.cell(3, col_idx).value for col_idx in range(1, ws.max_column + 1)]
        assert "Testnachweis" in headers

        sheet_ids = [
            str(ws.cell(row_idx, 1).value)
            for row_idx in range(4, ws.max_row + 1)
            if ws.cell(row_idx, 1).value
        ]
        expected_ids = set(df.loc[df["Metrik-ID"].str.startswith(f"{prefix}-"), "Metrik-ID"])
        assert expected_ids.issubset(set(sheet_ids))
        assert len(sheet_ids) == len(set(sheet_ids))


def test_analysis_glossary_review_documents_all_decision_findings():
    text = REVIEW_PATH.read_text(encoding="utf-8")

    for finding_id in range(1, 10):
        assert f"### F{finding_id} -" in text

    for phrase in [
        "Jobgruppe vs. Jobfamily/Cluster",
        "Delta % bei IST 0",
        "Unternehmensabgaenge mit negativem Headcount-Effekt",
        "Review-Entscheidungsvorschlag",
        "fachlich bestaetigt",
    ]:
        assert phrase in text
