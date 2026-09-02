"""
Datenintegritätsprüfung zwischen Mitarbeiter.xlsx und Planstellen.xlsx.

Kontext (Blocker B17): Das Dashboard geht davon aus, dass Mitarbeiter.xlsx
(Personal-Stammdaten) und Planstellen.xlsx (Stellenplan) deckungsgleich sind -
jede besetzte Planstelle hat genau einen passenden Mitarbeiter-Datensatz.
Ist das nicht der Fall, ist das ein Fehler im Datenlieferungsprozess, keine
Eigenschaft der echten Belegschaft. Das Dashboard schließt betroffene Zeilen
dann sauber aus den personenbezogenen KPIs aus (siehe
kpi_engine.get_unique_employees), statt sie falsch zu zählen - aber diese
Abweichungen müssen für den Nutzer sichtbar werden, sonst bleiben sie
unbemerkt. Dieses Modul stellt die Prüfung dafür bereit.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from abgaenge.schemas import normalize_persnr

_FORMULA_LEADING_CHARS = ("=", "+", "-", "@")


def _sanitize_excel_cell(value: Any) -> Any:
    """Verhindert Formel-Injection: Zellen, deren Text mit =, +, -, @ beginnt,
    würden von Excel/openpyxl sonst als Formel statt als Text interpretiert
    (kann aus einer manipulierten Upload-Datei stammen, die hier weiterverarbeitet
    und zum Download angeboten wird)."""
    if isinstance(value, str) and value[:1] in _FORMULA_LEADING_CHARS:
        return " " + value
    return value


def _now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


@dataclass
class IntegrityCheckResult:
    key: str
    title: str
    severity: str  # "error" | "warning"
    description: str
    detail: pd.DataFrame = field(default_factory=pd.DataFrame)

    @property
    def count(self) -> int:
        return len(self.detail)


@dataclass
class DataIntegrityReport:
    checks: list[IntegrityCheckResult] = field(default_factory=list)
    generated_at: str = ""
    mitarbeiter_row_count: int = 0
    planstellen_row_count: int = 0

    @property
    def total_findings(self) -> int:
        return sum(c.count for c in self.checks)

    @property
    def is_clean(self) -> bool:
        return self.total_findings == 0

    @property
    def error_count(self) -> int:
        return sum(c.count for c in self.checks if c.severity == "error")

    @property
    def warning_count(self) -> int:
        return sum(c.count for c in self.checks if c.severity == "warning")


def check_mitarbeiter_planstellen_integrity(
    mitarbeiter: pd.DataFrame,
    planstellen: pd.DataFrame,
) -> DataIntegrityReport:
    """
    Vergleicht Personalnummern zwischen Mitarbeiter.xlsx (Stammdaten, Spalte
    'PersNr') und Planstellen.xlsx (Stellenplan, Spalte 'Personalnummer').

    Drei Prüfungen:
    1. Besetzte Planstelle ohne Mitarbeiter-Match (Ursache von Blocker B17:
       diese Personen fallen im Kompakt-Headcount sauber heraus, sind aber
       echte, besetzte Stellen).
    2. Mitarbeiter ohne zugeordnete (besetzte) Planstelle.
    3. Doppelte Personalnummer innerhalb Mitarbeiter.xlsx (Stammdatenfehler,
       macht Zuordnungen uneindeutig).
    """
    ma = mitarbeiter.copy() if mitarbeiter is not None else pd.DataFrame()
    pl = planstellen.copy() if planstellen is not None else pd.DataFrame()

    checks: list[IntegrityCheckResult] = []

    if "PersNr" not in ma.columns or "Personalnummer" not in pl.columns:
        missing = []
        if "PersNr" not in ma.columns:
            missing.append("Mitarbeiter.xlsx: Spalte 'PersNr' fehlt")
        if "Personalnummer" not in pl.columns:
            missing.append("Planstellen.xlsx: Spalte 'Personalnummer' fehlt")
        checks.append(IntegrityCheckResult(
            key="schema_unvollstaendig",
            title="Erwartete Spalten fehlen",
            severity="error",
            description=(
                "Ohne diese Spalten kann die Datenintegrität zwischen Mitarbeiter.xlsx "
                "und Planstellen.xlsx nicht geprüft werden."
            ),
            detail=pd.DataFrame({"Fehlende Spalte": missing}),
        ))
        return DataIntegrityReport(
            checks=checks,
            generated_at=_now_iso(),
            mitarbeiter_row_count=len(ma),
            planstellen_row_count=len(pl),
        )

    ma["PersNr"] = normalize_persnr(ma["PersNr"])
    pl["Personalnummer"] = normalize_persnr(pl["Personalnummer"])

    besetzt = pl[pl["Personalnummer"].notna()].copy()
    ma_valid = ma[ma["PersNr"].notna()].copy()

    ma_persnr_set = set(ma_valid["PersNr"])
    pl_persnr_set = set(besetzt["Personalnummer"])

    # --- Check 1: besetzte Planstelle ohne Mitarbeiter-Match (B17-Fall) ---
    orphan_planstellen = besetzt[~besetzt["Personalnummer"].isin(ma_persnr_set)]
    orphan_pl_cols = [
        c for c in ("Personalnummer", "Planstellennr", "Kürzel OrgEinheit", "Organisationseinheit", "Planstelle")
        if c in orphan_planstellen.columns
    ]
    checks.append(IntegrityCheckResult(
        key="planstelle_ohne_mitarbeiter",
        title="Besetzte Planstelle ohne Mitarbeiter-Datensatz",
        severity="error",
        description=(
            "Diese Planstellen sind laut Planstellen.xlsx besetzt (Personalnummer vorhanden), "
            "aber die Personalnummer kommt in Mitarbeiter.xlsx nicht vor. Das Dashboard kann diese "
            "Personen keiner Kennzahl zuordnen und schließt sie sauber aus dem Headcount aus "
            "(z. B. Kompakt 'Gesamt Köpfe'), statt sie fehlerhaft mitzuzählen. Prüfen Sie, ob in "
            "Mitarbeiter.xlsx Datensätze fehlen oder die Personalnummer abweicht."
        ),
        detail=orphan_planstellen[orphan_pl_cols].reset_index(drop=True),
    ))

    # --- Check 2: Mitarbeiter ohne zugeordnete (besetzte) Planstelle ---
    orphan_mitarbeiter = ma_valid[~ma_valid["PersNr"].isin(pl_persnr_set)]
    orphan_ma_cols = [
        c for c in ("PersNr", "Vorname", "Nachname", "MitarbGruppenbez.", "Status kundenindividuell", "Eintritt", "Austritt")
        if c in orphan_mitarbeiter.columns
    ]
    checks.append(IntegrityCheckResult(
        key="mitarbeiter_ohne_planstelle",
        title="Mitarbeiter ohne zugeordnete Planstelle",
        severity="warning",
        description=(
            "Diese Personalnummern kommen in Mitarbeiter.xlsx vor, sind aber auf keiner besetzten "
            "Stelle in Planstellen.xlsx zu finden. Mögliche Ursachen: Austritt nicht erfasst, "
            "Planstellen.xlsx unvollständig, oder ein Tippfehler in der Personalnummer."
        ),
        detail=orphan_mitarbeiter[orphan_ma_cols].reset_index(drop=True),
    ))

    # --- Check 3: doppelte Personalnummer in Mitarbeiter.xlsx ---
    dupe_mask = ma_valid["PersNr"].duplicated(keep=False)
    dupes = ma_valid[dupe_mask].sort_values("PersNr")
    dupe_cols = [c for c in ("PersNr", "Vorname", "Nachname", "Eintritt", "Austritt") if c in dupes.columns]
    checks.append(IntegrityCheckResult(
        key="doppelte_personalnummer_mitarbeiter",
        title="Doppelte Personalnummer in Mitarbeiter.xlsx",
        severity="error",
        description=(
            "Mitarbeiter.xlsx ist die Stammdatei und sollte pro Person genau einen Datensatz "
            "enthalten. Diese Personalnummern kommen mehrfach vor, wodurch Zuordnungen im "
            "Dashboard uneindeutig werden."
        ),
        detail=dupes[dupe_cols].reset_index(drop=True),
    ))

    return DataIntegrityReport(
        checks=checks,
        generated_at=_now_iso(),
        mitarbeiter_row_count=len(ma),
        planstellen_row_count=len(pl),
    )


def build_integrity_report_excel(report: DataIntegrityReport) -> bytes:
    """Baut eine Evaluations-Excel mit einer Übersicht + einem Detail-Sheet
    je Prüfung. Alle Zellwerte laufen durch _sanitize_excel_cell, da die
    Detaildaten aus einer vom Nutzer hochgeladenen Datei stammen."""
    wb = Workbook()
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    overview = wb.active
    overview.title = "Übersicht"
    overview.append(["Datenintegritäts-Evaluation"])
    overview["A1"].font = Font(bold=True, size=14)
    overview.append(["Erstellt am (UTC)", report.generated_at])
    overview.append(["Mitarbeiter.xlsx – Zeilen", report.mitarbeiter_row_count])
    overview.append(["Planstellen.xlsx – Zeilen", report.planstellen_row_count])
    overview.append([])
    overview.append(["Prüfung", "Schweregrad", "Anzahl Treffer", "Beschreibung"])
    header_row = overview.max_row
    for col in range(1, 5):
        cell = overview.cell(row=header_row, column=col)
        cell.font = header_font
        cell.fill = header_fill

    for check in report.checks:
        overview.append([
            _sanitize_excel_cell(check.title),
            "Fehler" if check.severity == "error" else "Warnung",
            check.count,
            _sanitize_excel_cell(check.description),
        ])

    overview.column_dimensions["A"].width = 42
    overview.column_dimensions["B"].width = 12
    overview.column_dimensions["C"].width = 16
    overview.column_dimensions["D"].width = 90

    for check in report.checks:
        if check.detail.empty:
            continue
        sheet_title = check.title[:31] or check.key[:31]
        # Excel-Sheetnamen: max 31 Zeichen, keine Duplikate
        base_title = sheet_title
        suffix = 1
        while sheet_title in wb.sheetnames:
            suffix += 1
            sheet_title = f"{base_title[:28]}_{suffix}"
        ws = wb.create_sheet(title=sheet_title)

        columns = list(check.detail.columns)
        ws.append(columns)
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        for _, row in check.detail.iterrows():
            ws.append([_sanitize_excel_cell(_stringify_cell(v)) for v in row.tolist()])

        for col_idx, col_name in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(40, len(str(col_name)) + 6))

    buffer = _write_workbook_to_bytes(wb)
    return buffer


def _stringify_cell(value: Any) -> Any:
    if pd.isna(value):
        return ""
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d")
    return value


def _write_workbook_to_bytes(wb: Workbook) -> bytes:
    import io

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
